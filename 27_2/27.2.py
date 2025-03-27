import cgi
from string import Template

def convert_currency(file_path, amount, from_currency, to_currency):
    from openpyxl import load_workbook
    workbook = load_workbook(file_path)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        if row[0] == from_currency.upper() and row[1] == to_currency.upper():
            return amount * float(str(row[2]).replace(",", "."))
    return None

def application(environ, start_response):
    if environ.get("PATH_INFO", "").lstrip("/") == "":
        form = cgi.FieldStorage(fp=environ["wsgi.input"], environ=environ)
        amount = form.getfirst("amount", "").strip()
        from_currency = form.getfirst("from_currency", "").strip()
        to_currency = form.getfirst("to_currency", "").strip()

        if amount and from_currency and to_currency:
            try:
                amount = float(amount)
                file_path = "C:\\27_2\\course.xlsx"
                converted_amount = convert_currency(file_path, amount, from_currency, to_currency)
                if converted_amount is not None:
                    result = f"<h1>{amount} {from_currency.upper()} = {converted_amount:.2f} {to_currency.upper()}</h1>"
                else:
                    result = "<h1>Помилка: Курс для обраних валют не знайдено.</h1>"
            except ValueError:
                result = "<h1>Помилка: Будь ласка, введіть коректну суму.</h1>"
        else:
            result = "<h1>Помилка: Заповніть всі поля.</h1>"

        start_response("200 OK", [("Content-type", "text/html; charset=utf-8")])
        with open("template/course.html", encoding="utf-8") as f:
            content = Template(f.read()).substitute(result=result)
    else:
        start_response("404 NOT FOUND", [("Content-type", "text/html; charset=utf-8")])
        with open("template/error_404.html", encoding="utf-8") as f:
            content = f.read()
    return [bytes(content, encoding="utf-8")]

HOST = ""
PORT = 8000

if __name__ == "__main__":
    from wsgiref.simple_server import make_server
    httpd = make_server(HOST, PORT, application)
    print(f"Local webserver is running at http://localhost:{PORT}")
    httpd.serve_forever()
